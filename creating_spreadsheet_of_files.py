from pathlib import Path
from openpyxl import Workbook


def save_dirs_files(dirs, name_list, path_list):
    for file in dirs.iterdir():
        if file.is_dir():
            for child_file in file.iterdir():
                if child_file.is_dir():
                    path_list.append(str(child_file))
                    name_list.append(f'DIR {child_file.stem}')
                else:
                    path_list.append(str(child_file))
                    name_list.append(f'ARQ {child_file.stem}')
            save_dirs_files(file, name_list, path_list)
        else:
            path_list.append(str(file.parent))
            name_list.append(f'ARQ {file.stem}')


def list_files_spreadsheet(path, file_name):
    my_files = Path(path)
    path_list = []
    name_list = []
    work_book = Workbook()
    row, col1_w, col2_w = 0, 0, 0
    work_sheet = work_book.active

    save_dirs_files(my_files, name_list, path_list)

    while row <= len(name_list):
        name = work_sheet.cell(row=row + 1, column=1)
        path = work_sheet.cell(row=row + 1, column=2)

        if row == 0:
            name.value = "Nome"
            path.value = "Diretório"
            row += 1
            continue

        name.value = name_list[row - 1]
        path.value = path_list[row - 1]
        row += 1

    work_book.save(f"{file_name}.xlsx")


list_files_spreadsheet("\\Users\\lfanj\\Desktop\\Creations", "arquivos2")
